from __future__ import annotations

import csv
import io
import json
from collections import Counter
from dataclasses import dataclass
from itertools import zip_longest
from pathlib import Path

from workpytools.common.textfile import read_text_with_fallback

DEFAULT_EMPTY_VALUES = frozenset(
    {None, "", "N/A", "NULL", "null", "None", "none", "-", "NaN"}
)

_BOOL_TAG_PREFIX = "\x00bool\x00"


@dataclass(frozen=True)
class Table:
    """A single sheet/file worth of tabular data: column names plus row-major values."""

    columns: list[str]
    rows: list[tuple[object, ...]]


@dataclass(frozen=True)
class ColumnProfile:
    column: str
    rows: int
    filled: int
    empty: int
    unique: int
    is_filled: bool
    is_unique: bool
    fill_rate: float
    unique_rate: float
    key_score: float
    top: list[tuple[str, int]]


def normalize_value(value: object) -> object:
    """Fold Excel's int/float ambiguity (1 vs 1.0) and trim incidental whitespace.

    Values are compared as strings elsewhere, but this keeps `1` and `1.0`
    (a common artifact of Excel cell formatting) from being counted as
    distinct values. `bool` is tagged so `True` (the value) and `"True"`
    (the string) don't collide once both are stringified.
    """
    if isinstance(value, bool):
        return f"{_BOOL_TAG_PREFIX}{value}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, str):
        return value.strip()
    if value is None:
        return None
    return str(value)


def is_empty(value: object, empty_values: frozenset[object]) -> bool:
    if value in empty_values:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def profile_columns(
    table: Table,
    top_n: int = 10,
    empty_values: frozenset[object] = DEFAULT_EMPTY_VALUES,
) -> list[ColumnProfile]:
    """Profile every column of `table`.

    Rows are transposed with `zip_longest` before counting (measured faster
    than a single pass that updates one Counter per column per row: see
    doc/profiler.md for the benchmark). Ragged rows are padded with None by
    `zip_longest`, so uneven row lengths don't misalign columns.
    """
    n_rows = len(table.rows)
    results: list[ColumnProfile] = []

    if n_rows == 0:
        columns_by_index: list[tuple[int, tuple[object, ...]]] = [
            (i, ()) for i in range(len(table.columns))
        ]
    else:
        columns_by_index = list(enumerate(zip_longest(*table.rows)))

    for col_idx, raw_values in columns_by_index:
        name = table.columns[col_idx] if col_idx < len(table.columns) else str(col_idx)
        normalized = [normalize_value(v) for v in raw_values]
        counter = Counter(v for v in normalized if not is_empty(v, empty_values))

        filled = sum(counter.values())
        empty = n_rows - filled
        unique = len(counter)

        fill_rate = filled / n_rows if n_rows else 0.0
        unique_rate = unique / n_rows if n_rows else 0.0
        key_score = _harmonic_mean(fill_rate, unique_rate)

        top = counter.most_common(top_n) if top_n > 0 else []

        results.append(
            ColumnProfile(
                column=name,
                rows=n_rows,
                filled=filled,
                empty=empty,
                unique=unique,
                is_filled=filled == n_rows,
                is_unique=unique == n_rows,
                fill_rate=fill_rate,
                unique_rate=unique_rate,
                key_score=key_score,
                top=[(_display_value(v), c) for v, c in top],
            )
        )

    return results


def _display_value(value: object) -> str:
    """Undo normalize_value's internal bool tagging for display purposes."""
    if isinstance(value, str) and value.startswith(_BOOL_TAG_PREFIX):
        return value[len(_BOOL_TAG_PREFIX) :]
    return str(value)


def _harmonic_mean(a: float, b: float) -> float:
    if a <= 0 or b <= 0:
        return 0.0
    return 2 * a * b / (a + b)


def format_top(top: list[tuple[str, int]]) -> str:
    """`値(件数) / 値(件数)`, space-padded around `/` to reduce ambiguity
    with slashes that appear inside values (dates, paths).
    """
    return " / ".join(f"{value}({count})" for value, count in top)


# --- 入力の読み込み ---


def read_csv_like(text: str, sep: str, header_row: int | None) -> Table:
    reader = csv.reader(io.StringIO(text), delimiter=sep)
    all_rows: list[tuple[object, ...]] = [tuple(row) for row in reader if row]

    if header_row is None:
        columns = [str(i) for i in range(len(all_rows[0]))] if all_rows else []
        data_rows = all_rows
    else:
        columns = [str(c) for c in all_rows[header_row]] if header_row < len(all_rows) else []
        data_rows = all_rows[header_row + 1 :]

    return Table(columns=columns, rows=data_rows)


def read_json_records(text: str) -> Table:
    """Accepts either a JSON array of objects, or JSON Lines (one object per line).

    Nested values (dict/list) are kept as JSON strings rather than flattened;
    this is a flat-table profiler, not a schema explorer.
    """
    stripped = text.strip()
    if stripped.startswith("["):
        records = json.loads(stripped)
    else:
        records = [json.loads(line) for line in stripped.splitlines() if line.strip()]

    columns: list[str] = []
    seen_columns: set[str] = set()
    for record in records:
        for key in record:
            if key not in seen_columns:
                seen_columns.add(key)
                columns.append(key)

    rows = []
    for record in records:
        row = []
        for col in columns:
            value = record.get(col)
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            row.append(value)
        rows.append(tuple(row))

    return Table(columns=columns, rows=rows)


def read_excel_tables(path: Path, header_row: int | None) -> dict[str, Table]:
    """Read every sheet, read_only for speed (measured ~27% faster than the
    default mode; see doc/profiler.md). Excel is far slower to read than CSV
    of the same size regardless, so this only narrows the gap.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        result: dict[str, Table] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            all_rows = [row for row in worksheet.iter_rows(values_only=True) if row]

            if header_row is None:
                columns = [str(i) for i in range(len(all_rows[0]))] if all_rows else []
                data_rows = all_rows
            else:
                columns = (
                    [str(c) if c is not None else "" for c in all_rows[header_row]]
                    if header_row < len(all_rows)
                    else []
                )
                data_rows = all_rows[header_row + 1 :]

            result[sheet_name] = Table(columns=columns, rows=data_rows)
        return result
    finally:
        workbook.close()


SUPPORTED_SUFFIXES = frozenset({".csv", ".tsv", ".txt", ".xlsx", ".json", ".jsonl"})


def load_tables(path: Path, sep: str | None, header_row: int | None) -> dict[str, Table]:
    """Load every table in `path`. The dict key is the sheet name for Excel,
    or "" for everything else (single-table formats).
    """
    suffix = path.suffix.lower()

    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError(f"対応していない形式です: {path}")

    if suffix == ".xlsx":
        return read_excel_tables(path, header_row)

    text = read_text_with_fallback(path)

    if suffix in (".json", ".jsonl"):
        return {"": read_json_records(text)}

    resolved_sep = sep if sep is not None else ("\t" if suffix == ".tsv" else ",")
    return {"": read_csv_like(text, resolved_sep, header_row)}
