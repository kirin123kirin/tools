import argparse
from pathlib import Path

import pytest

from tools.processing import profiler as profiler_module
from tools.processing.profiler import ProfilerProcessor


def _base_args(**overrides: object) -> argparse.Namespace:
    defaults = dict(
        path=[],
        output=None,
        clip=False,
        view=False,
        sep=None,
        header=0,
        no_header=False,
        top=10,
        empty_values=None,
        no_default_empty_values=False,
    )
    defaults.update(overrides)
    return argparse.Namespace(**defaults)


@pytest.fixture
def clipboard_state(monkeypatch: pytest.MonkeyPatch) -> dict:
    state = {"text": None, "written": None}
    monkeypatch.setattr(profiler_module, "get_clipboard_text", lambda: state["text"])

    def fake_copy(text: str) -> None:
        state["written"] = text

    monkeypatch.setattr(profiler_module, "copy_text_to_clipboard", fake_copy)
    return state


def _write_csv(tmp_path: Path, name: str, content: str) -> Path:
    path = tmp_path / name
    path.write_text(content, encoding="utf-8")
    return path


# --- 入力: ファイル ---


def test_csv_file_profiled(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    csv_path = _write_csv(tmp_path, "a.csv", "a,b\n1,x\n2,y\n1,x\n")

    ProfilerProcessor().run(_base_args(path=[str(csv_path)]))

    out = capsys.readouterr().out
    lines = out.strip().split("\n")
    assert lines[0].split("\t")[0] == "source"
    assert "sheet" not in lines[0]  # Excelを含まないので出ない


def test_multiple_files_distinguished_by_source(
    tmp_path: Path, capsys: pytest.CaptureFixture
) -> None:
    a = _write_csv(tmp_path, "a.csv", "x\n1\n")
    b = _write_csv(tmp_path, "b.csv", "x\n2\n")

    ProfilerProcessor().run(_base_args(path=[str(a), str(b)]))

    out = capsys.readouterr().out
    assert str(a) in out
    assert str(b) in out


def test_header_only_once_across_files(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    a = _write_csv(tmp_path, "a.csv", "x\n1\n")
    b = _write_csv(tmp_path, "b.csv", "x\n2\n")

    ProfilerProcessor().run(_base_args(path=[str(a), str(b)]))

    out = capsys.readouterr().out
    assert out.count("source\t") == 1


def test_nonexistent_file_raises(tmp_path: Path) -> None:
    with pytest.raises(SystemExit):
        ProfilerProcessor().run(_base_args(path=[str(tmp_path / "missing.csv")]))


def test_unsupported_format_raises(tmp_path: Path) -> None:
    path = tmp_path / "a.pptx"
    path.write_bytes(b"dummy")
    with pytest.raises(SystemExit):
        ProfilerProcessor().run(_base_args(path=[str(path)]))


def test_empty_file_raises(tmp_path: Path) -> None:
    path = _write_csv(tmp_path, "empty.csv", "")
    with pytest.raises(SystemExit):
        ProfilerProcessor().run(_base_args(path=[str(path)]))


def test_one_empty_file_among_many_is_skipped_with_warning(
    tmp_path: Path, capsys: pytest.CaptureFixture, caplog: pytest.LogCaptureFixture
) -> None:
    good = _write_csv(tmp_path, "good.csv", "x\n1\n")
    empty = _write_csv(tmp_path, "empty.csv", "")

    with caplog.at_level("WARNING"):
        ProfilerProcessor().run(_base_args(path=[str(good), str(empty)]))

    out = capsys.readouterr().out
    assert str(good) in out
    assert "スキップ" in caplog.text


def test_no_header_option(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    path = _write_csv(tmp_path, "a.csv", "1,x\n2,y\n")

    ProfilerProcessor().run(_base_args(path=[str(path)], no_header=True))

    out = capsys.readouterr().out
    assert "\t0\t" in out or out.strip().split("\n")[1].split("\t")[2] == "0"


def test_sep_option_overrides_extension(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    path = _write_csv(tmp_path, "a.csv", "a;b\n1;x\n")

    ProfilerProcessor().run(_base_args(path=[str(path)], sep=";"))

    out = capsys.readouterr().out
    assert "a" in out and "b" in out


# --- 入力: Excel ---


def test_excel_multiple_sheets_produce_separate_records(
    tmp_path: Path, capsys: pytest.CaptureFixture
) -> None:
    import openpyxl

    path = tmp_path / "a.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "S1"
    ws1.append(["a"])
    ws1.append([1])
    ws2 = wb.create_sheet("S2")
    ws2.append(["b"])
    ws2.append([2])
    wb.save(path)

    ProfilerProcessor().run(_base_args(path=[str(path)]))

    out = capsys.readouterr().out
    assert "S1" in out
    assert "S2" in out


def test_sheet_column_present_when_excel_included(
    tmp_path: Path, capsys: pytest.CaptureFixture
) -> None:
    import openpyxl

    path = tmp_path / "a.xlsx"
    wb = openpyxl.Workbook()
    wb.active.append(["a"])
    wb.active.append([1])
    wb.save(path)

    ProfilerProcessor().run(_base_args(path=[str(path)]))

    out = capsys.readouterr().out
    header = out.strip().split("\n")[0]
    assert "sheet" in header.split("\t")


def test_sheet_column_absent_without_excel(
    tmp_path: Path, capsys: pytest.CaptureFixture
) -> None:
    path = _write_csv(tmp_path, "a.csv", "a\n1\n")

    ProfilerProcessor().run(_base_args(path=[str(path)]))

    out = capsys.readouterr().out
    header = out.strip().split("\n")[0]
    assert "sheet" not in header.split("\t")


# --- 入力: JSON ---


def test_json_array_of_objects(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    path = tmp_path / "a.json"
    path.write_text('[{"a": 1}, {"a": 2}]', encoding="utf-8")

    ProfilerProcessor().run(_base_args(path=[str(path)]))

    out = capsys.readouterr().out
    assert "a" in out


def test_jsonl(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    path = tmp_path / "a.jsonl"
    path.write_text('{"a": 1}\n{"a": 2}\n', encoding="utf-8")

    ProfilerProcessor().run(_base_args(path=[str(path)]))

    out = capsys.readouterr().out
    assert "a" in out


# --- 入力: クリップボード ---


def test_no_args_reads_clipboard(clipboard_state, capsys: pytest.CaptureFixture) -> None:
    clipboard_state["text"] = "a\tb\n1\tx\n2\ty\n"

    ProfilerProcessor().run(_base_args(path=[]))

    out = capsys.readouterr().out
    assert "(clipboard)" in out


def test_clipboard_empty_raises(clipboard_state) -> None:
    clipboard_state["text"] = None
    with pytest.raises(SystemExit):
        ProfilerProcessor().run(_base_args(path=[]))


# --- 出力 ---


def test_default_output_is_stdout_tsv(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    path = _write_csv(tmp_path, "a.csv", "a\n1\n")

    ProfilerProcessor().run(_base_args(path=[str(path)]))

    out = capsys.readouterr().out
    assert "\t" in out.split("\n")[0]


def test_output_tsv_file(tmp_path: Path) -> None:
    src = _write_csv(tmp_path, "a.csv", "a\n1\n")
    out_path = tmp_path / "result.tsv"

    ProfilerProcessor().run(_base_args(path=[str(src)], output=str(out_path)))

    assert out_path.exists()
    assert "source" in out_path.read_text(encoding="utf-8")


def test_output_xlsx_file(tmp_path: Path) -> None:
    src = _write_csv(tmp_path, "a.csv", "a\n1\n2\n")
    out_path = tmp_path / "result.xlsx"

    ProfilerProcessor().run(_base_args(path=[str(src)], output=str(out_path)))

    assert out_path.exists()

    import openpyxl

    wb = openpyxl.load_workbook(out_path)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][0] == "source"


def test_clip_option_copies_and_suppresses_stdout(
    tmp_path: Path, clipboard_state, capsys: pytest.CaptureFixture
) -> None:
    src = _write_csv(tmp_path, "a.csv", "a\n1\n")

    ProfilerProcessor().run(_base_args(path=[str(src)], clip=True))

    out = capsys.readouterr().out
    assert clipboard_state["written"] is not None
    assert "source\t" not in out  # 結果本体は標準出力に出ない
    assert "コピー" in out


def test_view_option_writes_html_and_opens_browser(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    src = _write_csv(tmp_path, "a.csv", "a\n1\n")

    calls = []
    monkeypatch.setattr(
        "tools.common.browser_preview.tempfile.gettempdir", lambda: str(tmp_path)
    )
    monkeypatch.setattr(
        "tools.common.browser_preview.webbrowser.open", lambda url: calls.append(url)
    )

    ProfilerProcessor().run(_base_args(path=[str(src)], view=True))

    assert len(calls) == 1
    written = list(tmp_path.glob("*.html"))
    assert len(written) == 1


def test_no_openpyxl_import_for_csv_only(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    import sys

    src = _write_csv(tmp_path, "a.csv", "a\n1\n")
    sys.modules.pop("openpyxl", None)

    ProfilerProcessor().run(_base_args(path=[str(src)]))

    assert "openpyxl" not in sys.modules


# --- 空とみなす値 ---


def test_empty_values_option_replaces_default(
    tmp_path: Path, capsys: pytest.CaptureFixture
) -> None:
    path = _write_csv(tmp_path, "a.csv", "a\nN/A\ncustom\n")

    ProfilerProcessor().run(_base_args(path=[str(path)], empty_values="custom"))

    out = capsys.readouterr().out
    data_row = out.strip().split("\n")[1].split("\t")
    # filled列: N/Aはもう空とみなされないので2件とも埋まっている扱いだが、
    # customは空扱いになるので filled=1
    assert data_row[3] == "1"  # filled


def test_no_default_empty_values(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    path = _write_csv(tmp_path, "a.csv", "a\nN/A\nx\n")

    ProfilerProcessor().run(_base_args(path=[str(path)], no_default_empty_values=True))

    out = capsys.readouterr().out
    data_row = out.strip().split("\n")[1].split("\t")
    assert data_row[3] == "2"  # N/Aも埋まっている扱いになる


def test_top_option_limits_count(tmp_path: Path, capsys: pytest.CaptureFixture) -> None:
    rows = "\n".join(str(i) for i in range(20))
    path = _write_csv(tmp_path, "a.csv", f"a\n{rows}\n")

    ProfilerProcessor().run(_base_args(path=[str(path)], top=2))

    out = capsys.readouterr().out
    top_field = out.strip().split("\n")[1].split("\t")[-1]
    assert top_field.count("/") == 1  # 2件なら区切りは1つ
